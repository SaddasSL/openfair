"""Generate AS9102 Form 3 (Characteristic Accountability) as an Excel workbook."""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

THIN = Side(style="thin")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="D9D9D9")
HDR_FONT = Font(bold=True, size=10)


def designator(entry: dict) -> str:
    t = entry.get("type")
    if t == "gdt":
        return "GD&T"
    if t == "note":
        return "Note"
    return "Dimension"


def requirement(entry: dict) -> str:
    v = (entry.get("value") or "").strip()
    tol = (entry.get("tolerance") or "").strip()
    if not tol:
        return v
    # some GD&T entries carry the frame in both fields - never print it twice
    nv, nt = (v.replace(" ", "").replace("|", "").replace("Ø", "⌀"), tol.replace(" ", "").replace("|", "").replace("Ø", "⌀"))
    if nv in nt:
        return tol
    if nt in nv:
        return v
    return f"{v} {tol}"


def form3(json_path: str, part_number: str = "", part_name: str = "") -> Path:
    entries = json.loads(Path(json_path).read_text(encoding="utf-8"))
    ballooned = [e for e in entries if e.get("balloon")]
    ballooned.sort(key=lambda e: e["balloon"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Form 3"

    # header block
    ws.merge_cells("A1:H1")
    ws["A1"] = "AS9102 First Article Inspection - Form 3: Characteristic Accountability, Verification and Compatibility Evaluation"
    ws["A1"].font = Font(bold=True, size=11)
    labels = [("A2", "1. Part Number:", part_number),
              ("C2", "2. Part Name:", part_name),
              ("E2", "3. Serial Number:", ""),
              ("G2", "4. FAIR Identifier:", "")]
    for cell, label, value in labels:
        ws[cell] = label
        ws[cell].font = HDR_FONT
        col = chr(ord(cell[0]) + 1)
        ws[f"{col}2"] = value

    # column headers
    cols = ["5. Char No.", "6. Reference Location", "7. Characteristic Designator",
            "8. Requirement", "9. Results", "10. Designed / Qualified Tooling",
            "11. Nonconformance Number", "Additional Data / Comments"]
    widths = [10, 20, 22, 34, 16, 24, 24, 30]
    for i, (c, w) in enumerate(zip(cols, widths), start=1):
        cell = ws.cell(row=4, column=i, value=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[cell.column_letter].width = w

    # characteristic rows
    for r, e in enumerate(ballooned, start=5):
        row = [e["balloon"], e.get("zone") or "", designator(e), requirement(e),
               "", "", "", ""]
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BOX
            cell.alignment = Alignment(vertical="center", wrap_text=(c in (2, 4, 8)))

    out = Path(json_path).with_name(Path(json_path).stem.replace("_tiled", "") + "_form3.xlsx")
    wb.save(out)
    print(f"Form 3 with {len(ballooned)} characteristics -> {out}")
    return out


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: py -m src.reports.form3 <extracted_json> [part_number] [part_name]")
        sys.exit(1)
    form3(sys.argv[1],
          sys.argv[2] if len(sys.argv) > 2 else "",
          sys.argv[3] if len(sys.argv) > 3 else "")


